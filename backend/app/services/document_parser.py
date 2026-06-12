from typing import List
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook


class DocumentChunk(BaseModel):
    content: str
    metadata: dict


from pydantic import BaseModel


class DocumentChunk(BaseModel):
    content: str
    metadata: dict


def parse_docx(file_path: str) -> List[DocumentChunk]:
    doc = DocxDocument(file_path)
    chunks = []
    for i, para in enumerate(doc.paragraphs):
        if para.text.strip():
            chunks.append(DocumentChunk(
                content=para.text,
                metadata={"paragraph": i, "source": file_path}
            ))
    return chunks


def parse_excel(file_path: str) -> List[DocumentChunk]:
    wb = load_workbook(file_path, read_only=True)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_text = " | ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                chunks.append(DocumentChunk(
                    content=row_text,
                    metadata={"sheet": sheet_name, "row": row_idx, "source": file_path}
                ))
    wb.close()
    return chunks


def parse_web_url(url: str) -> List[DocumentChunk]:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    for script in soup(["script", "style"]):
        script.decompose()

    text = soup.get_text(separator="\n", strip=True)
    chunks = []
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]

    for i, para in enumerate(paragraphs):
        chunks.append(DocumentChunk(
            content=para,
            metadata={"paragraph": i, "source": url}
        ))
    return chunks


def parse_file(file_path: str) -> List[DocumentChunk]:
    ext = Path(file_path).suffix.lower()
    if ext == ".docx":
        return parse_docx(file_path)
    elif ext in [".xlsx", ".xls"]:
        return parse_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def chunk_text(chunks: List[DocumentChunk], chunk_size: int = 1000, overlap: int = 200) -> List[DocumentChunk]:
    result = []
    for chunk in chunks:
        if len(chunk.content) <= chunk_size:
            result.append(chunk)
        else:
            text = chunk.content
            start = 0
            while start < len(text):
                end = start + chunk_size
                result.append(DocumentChunk(
                    content=text[start:end],
                    metadata=chunk.metadata.copy()
                ))
                start = end - overlap
    return result
